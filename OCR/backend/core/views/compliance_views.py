import os
import uuid
from datetime import datetime

from django.db.models import Prefetch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.shortcuts import get_object_or_404

from backend import settings
from consts.responses import Responses
from core.models import Batch, Document, DocumentField, Field
from core.views import SITE_ID_ASCO
from utils.asco_reports import generate_pdf_from_asco, generate_docx_from_asco
from utils.mandatory_standards import mandatory_standard_list


class ComplianceView(APIView):
    """
    Run compliance checks on a batch of documents.
    Supported checks: hs_code, field_comparison
    """

    def post(self, request, format=None):
        batch_id = request.data.get("batch_id")
        checks = request.data.get("checks", [])

        if not checks:
            return Response({"detail": "No checks provided"}, status=status.HTTP_400_BAD_REQUEST)

        # Move hardcoded IDs to settings or constants
        HSCODE_FIELD_ID = 1
        WEIGHT_FIELD_IDS = [2, 3, 4, 5]
        SHIPPING_FIELD_IDS = [6, 7, 8, 9]

        result = []

        # Pre-fetch all field names
        all_field_ids = set([HSCODE_FIELD_ID] + WEIGHT_FIELD_IDS + SHIPPING_FIELD_IDS)
        field_names = Field.objects.filter(id__in=all_field_ids).values('id', 'name')
        field_name_map = {f['id']: f['name'] for f in field_names}

        def create_validation_table(field_ids, title, description):
            """Helper to create validation tables for weight/shipping checks"""
            table = [['Topic', 'Result', 'Description']]

            # Fetch all values for these fields in one query
            field_values = DocumentField.objects.filter(
                field_id__in=field_ids,
                document__batch_id=batch_id
            ).values('field_id', 'hitl_value')

            # Group values by field_id
            values_by_field = {}
            for item in field_values:
                field_id = item['field_id']
                if field_id not in values_by_field:
                    values_by_field[field_id] = []
                values_by_field[field_id].append(item['hitl_value'])

            # Create table rows
            for field_id in field_ids:
                field_name = field_name_map.get(field_id, f"Field {field_id}")
                values = values_by_field.get(field_id, [])

                if len(set(values)) > 1:
                    table.append([field_name, '⚠️ Incompatible', ', '.join(values)])
                elif len(set(values)) == 1:
                    table.append([field_name, '✅ Compatible', values[0]])

            return {
                'title': title,
                'description': description,
                'table': table
            }

        # Process checks
        for check in checks:
            if check == 'standard_validation':
                standard_table = [['HS Code', 'Requirement']]

                # Fetch all HS codes in one query
                hs_codes = DocumentField.objects.filter(
                    field_id=HSCODE_FIELD_ID,
                    document__batch_id=batch_id
                ).values_list('hitl_value', flat=True).distinct()

                # Check each HS code
                if not list(hs_codes):
                    standard_table.append(['', 'No HS Code exists.'])
                for hs_code in hs_codes:
                    if hs_code and hs_code in mandatory_standard_list:
                        standard_table.append([hs_code, 'Standard action required.'])

                # Only add if there are results
                if len(standard_table) > 1:
                    result.append({
                        'title': 'کالاهای استاندارد اجباری',
                        'description': 'لیست کالاهایی که دارای استاندارد اجباری می‌باشند...',
                        'table': standard_table,
                    })

            elif check == 'weight_quantity':
                result.append(create_validation_table(
                    field_ids=WEIGHT_FIELD_IDS,
                    title='اعتبارسنجی وزن و تعداد',
                    description='بررسی هم‌خوانی مشخصات مرتبط با وزن و تعداد...'
                ))

            elif check == 'shipping_validation':
                result.append(create_validation_table(
                    field_ids=SHIPPING_FIELD_IDS,
                    title='اعتبارسنجی اطلاعات حمل',
                    description='بررسی هم‌خوانی مشخصات مرتبط با اطلاعات حمل...'
                ))

        return Responses.get_response(Responses.GENERAL.OK, result)



class ReportView(APIView):
    def post(self, request, format=None):
        batch_id = request.data.get("batch_id")
        site_id = request.headers.get('X-Site-ID')

        if site_id == SITE_ID_ASCO:
            asco = request.data.get("asco")

            report_path, report_url = generate_report_path('pdf')
            pdf_url = generate_pdf_from_asco(asco, report_path, report_url)
            report_path, report_url = generate_report_path('docx')
            docx_url = generate_docx_from_asco(asco, report_path, report_url)

            return Responses.get_response(Responses.GENERAL.OK, {
                "pdf": pdf_url,
                "docx": docx_url
            })



        if not batch_id:
            return Responses.get_response(Responses.GENERAL.FORBIDDEN)

        # Fetch batch
        batch = get_object_or_404(Batch, id=batch_id)
        batch.status = Batch.STATUS_DONE
        batch.save()

        docs_count = Document.objects.filter(batch=batch).count()
        fields_count = DocumentField.objects.filter(document__batch=batch).count()

        # Generate Excel
        # Query documents based on filters
        query = Document.objects.filter(batch=batch)

        # Prefetch related data for performance
        documents = query.select_related('type', 'user', 'batch').prefetch_related(
            Prefetch(
                'documentfield_set',
                queryset=DocumentField.objects.select_related('field').order_by('field__group', 'field__name')
            )
        )

        if not documents.exists():
            raise ValueError("No documents found for the given criteria")

        # Generate unique filename
        random_secure_name = uuid.uuid4().hex[:16]
        timestamp = datetime.now().strftime("%Y%m%d%H%M")
        filename = f"compliance_{random_secure_name}_{timestamp}.xlsx"
        reports_dir = os.path.join(settings.BASE_DIR, 'static', 'reports')
        os.makedirs(reports_dir, exist_ok=True)
        report_path = os.path.join(reports_dir, filename)
        report_url = f"{settings.STATIC_URL}reports/{filename}".strip("/")

        # Ensure directory exists
        os.makedirs(os.path.dirname(report_path), exist_ok=True)

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)

        # Process each document
        for idx, document in enumerate(documents, 1):
            self._add_document_sheet(wb, document, idx)

        # Add summary sheet
        self._add_summary_sheet(wb, documents)

        # Save workbook
        wb.save(report_path)

        # Return success and report path
        return Responses.get_response(Responses.GENERAL.OK, {
            "report_path": report_url,
            "docs_count": docs_count,
            "fields_count": fields_count,
        })


    def _add_document_sheet(self, workbook, document, sheet_number):
        """Add a sheet for each document"""
        # Create sheet name (limit to 31 characters for Excel)
        sheet_name = f"Doc_{sheet_number}_{document.id}"[:31]
        ws = workbook.create_sheet(title=sheet_name)

        # Document header
        ws.merge_cells('A1:C1')
        title_cell = ws['A1']
        title_cell.value = f"Document ID: {document.id}"
        title_cell.font = Font(size=14, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Document metadata
        metadata = [
            ("Batch ID", document.batch.id if document.batch else "N/A"),
            ("Document Type", document.type.name if document.type else "N/A"),
            ("User", document.user.get_full_name() if document.user else document.user.username),
            ("Created",
             document.created_at.strftime("%Y-%m-%d %H:%M") if hasattr(document, 'created_at') else "N/A")
        ]

        for i, (key, value) in enumerate(metadata, start=3):
            ws[f"A{i}"] = key
            ws[f"A{i}"].font = Font(bold=True)
            ws[f"B{i}"] = value

        # Fields header
        ws['A7'] = "Group"
        ws['B7'] = "Field Name"
        ws['C7'] = "HITL Value"

        # Style header
        for col in ['A', 'B', 'C']:
            cell = ws[f"{col}7"]
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add field data
        row = 8
        current_group = None

        # Group fields by group and sort
        document_fields = document.documentfield_set.all()
        sorted_fields = sorted(document_fields, key=lambda x: (x.field.group or "", x.field.name))

        for doc_field in sorted_fields:
            field = doc_field.field

            # Add group row if group changes
            if field.group != current_group:
                if current_group is not None:
                    row += 1  # Add blank row between groups

                if field.group:
                    ws[f"A{row}"] = field.group
                    ws[f"A{row}"].font = Font(bold=True, italic=True, color="366092")
                    ws[f"B{row}"] = ""
                    ws[f"C{row}"] = ""
                    row += 1

                current_group = field.group

            # Add field data
            ws[f"A{row}"] = field.group or ""
            ws[f"B{row}"] = field.name
            ws[f"C{row}"] = doc_field.hitl_value or ""

            # Apply border
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            for col in ['A', 'B', 'C']:
                ws[f"{col}{row}"].border = thin_border

            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _add_summary_sheet(self, workbook, documents):
        """Add a summary sheet with overview of all documents"""
        ws = workbook.create_sheet(title="Summary")

        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = "Compliance Documents Summary"
        title_cell.font = Font(size=16, bold=True, color="FFFFFF")
        title_cell.alignment = Alignment(horizontal='center')
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Summary header
        headers = ["Document ID", "Batch ID", "Document Type", "User", "Fields Count", "Sheet"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')

        # Add document summaries
        for row_idx, document in enumerate(documents, start=4):
            ws.cell(row=row_idx, column=1, value=document.id)
            ws.cell(row=row_idx, column=2, value=document.batch.id if document.batch else "N/A")
            ws.cell(row=row_idx, column=3, value=document.type.name if document.type else "N/A")
            ws.cell(row=row_idx, column=4,
                    value=document.user.get_full_name() if document.user else document.user.username)
            ws.cell(row=row_idx, column=5, value=document.documentfield_set.count())
            ws.cell(row=row_idx, column=6, value=f"Doc_{row_idx - 3}_{document.id}"[:31])

            # Add hyperlink to sheet
            ws.cell(row=row_idx, column=6).hyperlink = f"#'Doc_{row_idx - 3}_{document.id}'!A1"
            ws.cell(row=row_idx, column=6).style = "Hyperlink"

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width



def generate_report_path(extension: str):
    random_secure_name = uuid.uuid4().hex[:16]
    timestamp = datetime.now().strftime("%Y%m%d%H%M")
    filename = f"asco_{random_secure_name}_{timestamp}.{extension}"

    reports_dir = os.path.join(settings.MEDIA_ROOT, "reports")
    os.makedirs(reports_dir, exist_ok=True)

    report_path = os.path.join(reports_dir, filename)

    # Public URL served by nginx
    report_url = f"{settings.MEDIA_URL}reports/{filename}"

    return report_path, report_url
